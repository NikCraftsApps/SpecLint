from __future__ import annotations
from typing import List, Dict, Tuple, Any
from pathlib import Path
from speclint.core.models import Requirement

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise RuntimeError("openpyxl is required for .xlsx parsing. Install package first.") from e


def _norm(s: str) -> str:
    """Normalize a header cell for matching: lowercase + keep only alphanumerics."""
    return "".join(ch for ch in s.lower() if ch.isalnum())


def _build_alias_map(cfg_cols: Dict[str, List[str]]) -> Dict[str, set[str]]:
    """Build a normalized alias set for each logical field."""
    alias_map: Dict[str, set[str]] = {}
    for logical, aliases in cfg_cols.items():
        alias_map[logical] = {_norm(a) for a in aliases}
    return alias_map


def _detect_header_row(ws, max_rows: int, alias_map: Dict[str, set[str]], required: set[str]) -> Tuple[int, Dict[str, int]]:
    """
    Scan the first `max_rows` rows to find a header row that contains all required fields.
    Returns (row_index_1based, mapping_logical_field -> column_index_1based).
    """
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        header_map: Dict[str, int] = {}
        for c_idx, val in enumerate(row or (), start=1):
            if val is None:
                continue
            key = _norm(str(val))
            if not key:
                continue
            for logical, keys in alias_map.items():
                if logical not in header_map and key in keys:
                    header_map[logical] = c_idx
        if required.issubset(header_map.keys()):
            return r_idx, header_map
    # If we got here: header not found
    miss = required - set(header_map.keys())
    raise ValueError(
        f"Could not detect header row with required columns {sorted(miss)} "
        f"(searched first {max_rows} rows)"
    )


def parse_xlsx_requirements(path: Path, cfg: Dict[str, Any]) -> List[Requirement]:
    """
    XLSX parser with flexible header/aliases controlled by config.
    Config path: inputs.xlsx, inputs.common.
    Required fields: id, title, risk. Optional: tests, tags.
    """
    xcfg = (cfg.get("inputs", {}) or {}).get("xlsx", {}) or {}
    common = (cfg.get("inputs", {}) or {}).get("common", {}) or {}
    sheet = xcfg.get("sheet", 0)
    max_hdr = int(xcfg.get("header_row_search_rows", 5))
    tests_sep = str(common.get("tests_separator", "|"))
    tags_sep = str(common.get("tags_separator", "|"))

    # column aliases (user can override defaults)
    default_cols = {
        "id":    ["id", "req id", "requirement id", "requirement"],
        "title": ["title", "name", "summary", "requirement title"],
        "risk":  ["risk", "severity", "priority"],
        "tests": ["tests", "test ids", "test cases", "test_cases"],
        "tags":  ["tags", "labels", "category"],
    }
    user_cols = xcfg.get("columns", {}) or {}
    for k, v in user_cols.items():
        default_cols[k] = list(dict.fromkeys(v))  # user-first, unique
    alias_map = _build_alias_map(default_cols)

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    # pick sheet by name or index
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        if sheet not in wb.sheetnames:
            raise ValueError(f"{path}: sheet '{sheet}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet]

    required = {"id", "title", "risk"}
    header_r, header_map = _detect_header_row(ws, max_hdr, alias_map, required)

    out: List[Requirement] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_r + 1, values_only=True), start=header_r + 1):
        if not row or not any(row):
            continue

        def get(logical: str) -> str:
            c = header_map.get(logical)
            if not c:
                return ""
            v = row[c - 1]
            return "" if v is None else str(v).strip()

        rid = get("id")
        title = get("title")
        risk = (get("risk") or "").lower() or None
        tests_raw = get("tests")
        tags_raw = get("tags")

        if not (rid or title or risk or tests_raw or tags_raw):
            continue

        tests = [t.strip() for t in (tests_raw.split(tests_sep) if tests_raw else []) if t.strip()]
        tags = [t.strip() for t in (tags_raw.split(tags_sep) if tags_raw else []) if t.strip()]

        out.append(
            Requirement(
                id=rid,
                title=title,
                risk=risk,
                tests=tests,
                tags=tags,
                file=str(path),
                line=r_idx,
            )
        )
    return out